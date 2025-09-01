import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import DataBar
from openpyxl.styles import Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBarRule


# Define the Excel file name
file_name = "Order_Analysis.xlsx"

# Categories for items
categories = {
    "Chairs": ["chair"],
    "Headsets": ["headset","phone"],
    "Smart Watches": ["smart watch", "smartwatch"],
    "Keyboards": ["keyboard"],
    "Laptops": ["laptop", "notebook"],
    "Monitors": ["monitor", "display", "screen"],
    "Gaming Tables": ["table", "desk"],
    "CPU Coolers": ["cpu cooler", "air cooler", "liquid cooler"],
    "PC Cases": ["pc case"],
    "PC Case Fans": ["pc case fan"],
    "Power Supplies": ["power supply"],
    "Power Banks": ["power bank"],
    "Bluetooth Speakers": ["speaker", "bt speaker"],
    "Wireless Earbuds": ["earbuds", "tws"],
    "Wireless Mice": ["wireless mouse", "gaming mouse"],
    "Combo": ["combo", "bundle", "workbuddy", "work"]  ,
    "Gift Wrapping":["Gift ","wrap"]
}

# Function to categorize items
def categorize_item(item):
    # First check if the item matches the "Combo" category
    if any(keyword.lower() in item.lower() for keyword in categories["Combo"]):
        return "Combo"
    
    # Then check other categories
    for category, keywords in categories.items():
        if category != "Combo" and any(keyword.lower() in item.lower() for keyword in keywords):
            return category
    return "Other"  # Return "Other" if no match


# Loading the file
dataframe = pd.read_csv("orders_export_1.csv")

# Exclude canceled orders (if 'cancelled_at' is not null)
dataframe = dataframe[dataframe['Cancelled at'].isnull()]

# Categorizing items
dataframe['Category'] = dataframe['Lineitem name'].apply(categorize_item)

# Grouping by city and counting the orders
OrderPerCity = dataframe.groupby('Billing City')['Lineitem quantity'].count()

# Identifying the city with the most orders
HighestOrderingCity = OrderPerCity.idxmax()  # City with the highest number of orders
OrdersOF_HOC = OrderPerCity.max()  # Number of orders that city has placed
print(f"The city with the most orders is {HighestOrderingCity} with {OrdersOF_HOC} orders.")

#=================================================================

# Filter data for the highest ordering city
city_data = dataframe[dataframe['Billing City'] == HighestOrderingCity]

# Grouping by category and counting the orders for that city
CategoryOrders = city_data.groupby('Category')['Lineitem quantity'].count()

# Saving results to an Excel file
CategoryOrders.to_frame(name="Order Count").reset_index().to_excel(
    f"Order_Analysis.xlsx", index=False, engine="openpyxl"
)

print("\nResults saved to Excel file:", f"Order_Analysis.xlsx")
print("\n\n\n\n")


#========================================

# Grouping by city and category, and counting the orders
OrderPerCityCategory = dataframe.groupby(['Billing City', 'Category'])['Lineitem quantity'].count()

# Reset the index to make it a DataFrame
OrderPerCityCategory_df = OrderPerCityCategory.reset_index(name="Order Count")

# Save to the same Excel file
file_name = "Order_Analysis.xlsx"  
with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    OrderPerCityCategory_df.to_excel(writer, sheet_name="Order Count", index=False)

print(f"\nResults saved to Excel file: {file_name}")

#-----------------------------------

# Group by Category and City, then sum the quantities
city_category_orders = dataframe.groupby(['Category', 'Billing City'])['Lineitem quantity'].sum()

# Convert the result to a DataFrame
city_category_df = city_category_orders.reset_index()

# Find the city with the highest orders for each category
top_city_per_category = city_category_df.loc[city_category_df.groupby('Category')['Lineitem quantity'].idxmax()]

# Rename columns for clarity
top_city_per_category.rename(columns={'Billing City': 'Top City', 'Lineitem quantity': 'Max Orders'}, inplace=True)

# Save to the same Excel file
with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    top_city_per_category.to_excel(writer, sheet_name="Top_Cities_Per_Category", index=False)
    
print(f"\nResults saved to Excel file: {file_name}")

#====================================

# Grouping by category and summing the item quantities for each category
CategoryItemCount = dataframe.groupby('Category')['Lineitem quantity'].sum()

# Grouping by category and product, and summing the item quantities for each product
ProductCountPerCategory = dataframe.groupby(['Category', 'Lineitem name'])['Lineitem quantity'].sum()

# Grouping by category to get the total item count per category
CategoryItemCount = ProductCountPerCategory.groupby('Category').sum()

# Saving results to an Excel file

# Step 1: Load existing Excel file without replacing it
wb = load_workbook(file_name)

# Step 2: Write new data (assuming CategoryItemCount and ProductCountPerCategory are DataFrames)
with pd.ExcelWriter(file_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    # Saving the new data (for example: 'Category Item Count' sheet)
    CategoryItemCount.to_frame(name="Total Item Count").reset_index().to_excel(writer, sheet_name="Category Item Count", index=False)
    ProductCountPerCategory.to_frame(name="Product Item Count").reset_index().to_excel(writer, sheet_name="Product Item Count", index=False)

# Step 3: Reopen the workbook to preserve formatting and apply manual formatting
wb = load_workbook(file_name)

# Access the sheet you want to format (example: "Category Item Count")
ws = wb["Category Item Count"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
   cell.font = Font(bold=True, size=12)   # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 30  # Adjust column 'B' width

# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border

ws = wb["Product Item Count"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
   cell.font = Font(bold=True, size=12)  # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 55  # Adjust column 'B' width
ws.column_dimensions['C'].width = 30 
# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border



 
 
ws = wb["Top_Cities_Per_Category"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
   cell.font = Font(bold=True, size=12)  # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 20  # Adjust column 'B' width
ws.column_dimensions['C'].width = 20 
# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border



ws = wb["Order Count"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
 cell.font = Font(bold=True, size=12) 
# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 50  # Adjust column 'B' width
ws.column_dimensions['C'].width = 20 
# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border


ws = wb["Top_Cities_Per_Category"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
    cell.font = Font(bold=True, size=12)  # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 20  # Adjust column 'B' width
ws.column_dimensions['C'].width = 20 
# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border

ws = wb["Sheet1"]

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in ws[1]:  # The first row is the header row
    cell.font = Font(bold=True, size=12)  # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
ws.column_dimensions['A'].width = 30  # Adjust column 'A' width
ws.column_dimensions['B'].width = 20  # Adjust column 'B' width

# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
from openpyxl.styles import Border, Side
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin_border



# Select the sheet where you want to apply table formatting
sheet = wb['Sheet1']  # You can change this to the sheet you want to format

# Define the range for your table (adjust range to your data's size)
# Example: 'A1:F100' includes your headers and data
table_range = 'A1:B17'

# Create a table object
table = Table(displayName="Highest_Ordering_City", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)


# Select the sheet where you want to apply table formatting
sheet = wb['Order Count']  # You can change this to the sheet you want to format

# Define the range for your table (adjust range to your data's size)
# Example: 'A1:F100' includes your headers and data
table_range = 'A1:C1271'

# Create a table object
table = Table(displayName="Order_Count", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)



# Select the sheet where you want to apply table formatting
sheet = wb['Top_Cities_Per_Category']  # You can change this to the sheet you want to format

# Define the range for your table (adjust range to your data's size)
# Example: 'A1:F100' includes your headers and data
table_range = 'A1:C16'

# Create a table object
table = Table(displayName="Top_Cities_Per_Category", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)



# Select the sheet where you want to apply table formatting
sheet = wb['Category Item Count']  # You can change this to the sheet you want to format

# Define the range for your table (adjust range to your data's size)
# Example: 'A1:F100' includes your headers and data
table_range = 'A1:B16'

# Create a table object
table = Table(displayName="Category_Item_Count", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)

# Select the sheet where you want to apply table formatting
sheet = wb['Product Item Count']  # You can change this to the sheet you want to format

# Define the range for your table (adjust range to your data's size)
# Example: 'A1:F100' includes your headers and data
table_range = 'A1:C107'

# Create a table object
table = Table(displayName="Product_Item_Count", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)


# Step 4: Save the workbook with the applied formatting and the new data
wb.save(file_name)

print(f"Data and formatting saved to '{file_name}'.")


#==================================================

#Calncels per city 

# Load the CSV file
file_path = "orders_export_1.csv"  # Update with your actual CSV file path
excel_file = "Order_Analysis.xlsx"  # Your existing Excel file

# Read CSV into DataFrame
df = pd.read_csv(file_path)

# Standardize city names (convert to lowercase and strip spaces)
df['Billing City'] = df['Billing City'].astype(str).str.strip().str.lower()

# Filter out rows where "Cancelled at" is not empty (meaning the order was canceled)
canceled_orders = df[df['Cancelled at'].notna()]

# Count the number of canceled orders per city
cancel_counts = canceled_orders.groupby('Billing City').size().reset_index(name='Cancel Count')

# Save to an existing Excel file without deleting other sheets
try:
    with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        cancel_counts.to_excel(writer, sheet_name="Cancelled Orders by City", index=False)
    print(f"Cancel counts saved to {excel_file} in 'Cancelled Orders by City' sheet.")
except FileNotFoundError:
    cancel_counts.to_excel(excel_file, sheet_name="Cancelled Orders by City", index=False)
    print(f"New file created: {excel_file} with 'Cancelled Orders by City' sheet.")

# Reopen the workbook after saving
wb = load_workbook(excel_file)

# Print available sheets to verify the correct name
print("Available sheets:", wb.sheetnames)

# Ensure correct sheet name
sheet_name = None
for name in wb.sheetnames:
    if "cancelled" in name.lower():
        sheet_name = name  # Assign the correct sheet name

if sheet_name is None:
    raise KeyError("Worksheet with 'Cancelled' in name not found!")

# Select the correct sheet
sheet = wb[sheet_name]

# Get the last row dynamically
last_row = sheet.max_row

# Define the range for your table dynamically
table_range = f"A1:B388"  # Adjust columns if necessary





# Create a table object
table = Table(displayName="CancelledOrdersByCity", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)

# Example: Reapply formatting manually (e.g., making headers bold)
for cell in sheet[1]:  # The first row is the header row
    cell.font = Font(bold=True, size=12)  # Set bold and font size 12

# Example: Adjusting column widths manually for better readability
sheet.column_dimensions['A'].width = 30  # Adjust column 'A' width
sheet.column_dimensions['B'].width = 20  # Adjust column 'B' width

# You can also add other formatting (colors, borders, etc.)
# Example: Adding a border to all cells in the sheet (you can customize this)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

for row in sheet.iter_rows():
    for cell in row:
        cell.border = thin_border





# Save and close
wb.save(excel_file)
wb.close()

print("Table formatting applied successfully!")










#--------------------


# Load existing workbook
file_name = "Order_Analysis.xlsx"
wb = load_workbook(file_name)

# Apply Data Bars to all sheets with relevant columns
for sheet in wb.worksheets:
    for col in sheet.iter_cols(min_col=2, max_col=sheet.max_column, min_row=2):  
        if all(isinstance(cell.value, (int, float)) for cell in col if cell.value is not None):  # Check if column has numbers
            rule = DataBarRule(start_type="num", start_value=min(cell.value for cell in col if cell.value is not None),
                               end_type="num", end_value=max(cell.value for cell in col if cell.value is not None),
                               color=Color("FF638EC6"))  # Blue color
            sheet.conditional_formatting.add(f"{col[0].column_letter}2:{col[0].column_letter}{sheet.max_row}", rule)

# Save changes
wb.save(file_name)
print("Data bars added successfully!")




#-----------------------------------


































#----------------------------------------------
###############################################

#Financial Metrics

# Ensure 'Total' column is numeric (handling potential errors)
dataframe['Total'] = pd.to_numeric(dataframe['Total'], errors='coerce')

# Calculate total revenue
total_revenue = dataframe['Total'].sum()

print(f'Total Revenue: ${total_revenue:.2f}')




# Calculate Total Orders (unique order IDs)
total_orders = dataframe["Id"].nunique()

# Calculate Average Order Value (AOV)
if total_orders > 0:
    aov = total_revenue / total_orders
else:
    aov = 0

print(f"Average Order Value (AOV): ${aov:.2f}")


# Ensure the relevant columns are treated as numeric
dataframe['Discount Amount'] = pd.to_numeric(dataframe['Discount Amount'], errors='coerce')
dataframe['Taxes'] = pd.to_numeric(dataframe['Taxes'], errors='coerce')

# Calculate total discounts and total taxes
total_discounts_given = dataframe['Discount Amount'].sum()
total_taxes_collected = dataframe['Taxes'].sum()

print(f"Total Discounts Given: {total_discounts_given}")
print(f"Total Taxes Collected: {total_taxes_collected}")

# Calculate total shipping revenue
total_shipping_revenue = dataframe['Shipping'].sum()
    
    # Calculate total revenue
total_revenue = dataframe['Total'].sum()
    
    # Calculate refund rate (avoid division by zero)
total_refunded = dataframe['Refunded Amount'].sum()
refund_rate = (total_refunded / total_revenue) * 100 if total_revenue > 0 else 0
    
    # Calculate total outstanding balance
total_outstanding_balance = dataframe['Outstanding Balance'].sum()
    
print(f'Total Shipping Revenue: {total_shipping_revenue}')
print(f'Refund Rate (%): {refund_rate}')

print(f'Total Outstanding Balance: {total_outstanding_balance}')

#--------------------------------------------

#Order & Fulfillment Metrics


# Convert date columns to datetime format
date_columns = ['Paid at', 'Fulfilled at', 'Cancelled at']
for col in date_columns:
    dataframe[col] = pd.to_datetime(dataframe[col], errors='coerce')  # Convert and handle errors


# Total Orders
total_orders = dataframe['Id'].nunique()

# Total Items Sold
total_items_sold = dataframe['Lineitem quantity'].sum()

# Average Items per Order
avg_items_per_order = total_items_sold / total_orders if total_orders > 0 else 0

# Fulfillment Rate
fulfilled_orders = dataframe['Fulfillment Status'].str.lower().eq('fulfilled').sum()
fulfillment_rate = (fulfilled_orders / total_orders) * 100 if total_orders > 0 else 0

# Time to Fulfillment (Average time in days)
fulfilled_df = dataframe.dropna(subset=['Paid at', 'Fulfilled at'])
fulfilled_df['Time to Fulfillment'] = (fulfilled_df['Fulfilled at'] - fulfilled_df['Paid at']).dt.total_seconds() / 86400
avg_time_to_fulfillment = fulfilled_df['Time to Fulfillment'].mean()

# Cancelled Orders
cancelled_orders = dataframe['Cancelled at'].notna().sum()


# Print results
print(f"Total Orders: {total_orders}")
print(f"Total Items Sold: {total_items_sold}")
print(f"Average Items per Order: {avg_items_per_order}")
print(f"Fulfillment Rate (%): {fulfillment_rate}")
print(f"Average Time to Fulfillment (days): {avg_time_to_fulfillment}")
print(f"Cancelled Orders: {cancelled_orders}")



#-----------------------------------------------

# 1. Best-Selling Products (Top Lineitem name based on Lineitem quantity)
best_selling_products = dataframe.groupby('Lineitem name')['Lineitem quantity'].sum().sort_values(ascending=False)

# 2. Revenue by Product (Sum of Lineitem price * Lineitem quantity per Lineitem name)
dataframe['Revenue'] = dataframe['Lineitem price'] * dataframe['Lineitem quantity']
revenue_by_product = dataframe.groupby('Lineitem name')['Revenue'].sum().sort_values(ascending=False)

# 3. Most Discounted Products (Products with the highest sum of Lineitem discount)
most_discounted_products = dataframe.groupby('Lineitem name')['Lineitem discount'].sum().sort_values(ascending=False)

# 4. Most Expensive Products Sold (Highest Lineitem price)
most_expensive_products = dataframe.groupby('Lineitem name')['Lineitem price'].max().sort_values(ascending=False)

# Display results
print("Best-Selling Products (Top by Quantity):")
print(best_selling_products)

print("\nRevenue by Product (Total Revenue):")
print(revenue_by_product)

print("\nMost Discounted Products (Total Discount Given):")
print(most_discounted_products)

print("\nMost Expensive Products Sold:")
print(most_expensive_products)




#--------------
# saving to excel 


# Prepare the Excel writer
with pd.ExcelWriter('Order_Analysis.xlsx', engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
 # Financial Metrics (Vertical Layout)
    financial_data = {
        'Metric': ['Total Revenue', 'Average Order Value (AOV)', 'Total Discounts Given', 
                   'Total Taxes Collected', 'Total Shipping Revenue', 'Refund Rate (%)', 
                   'Total Outstanding Balance'],
        'Value': [total_revenue, aov, total_discounts_given, total_taxes_collected, 
                  total_shipping_revenue, refund_rate, total_outstanding_balance]
    }
    financial_df = pd.DataFrame(financial_data)
    financial_df.to_excel(writer, sheet_name='Financial Metrics', index=False)

    # Order & Fulfillment Metrics (Vertical Layout)
    order_fulfillment_data = {
        'Metric': ['Total Orders', 'Total Items Sold', 'Average Items per Order', 'Fulfillment Rate (%)',
                   'Average Time to Fulfillment (days)', 'Cancelled Orders'],
        'Value': [total_orders, total_items_sold, avg_items_per_order, fulfillment_rate,
                  avg_time_to_fulfillment, cancelled_orders]
    }
    order_fulfillment_df = pd.DataFrame(order_fulfillment_data)
    order_fulfillment_df.to_excel(writer, sheet_name='Order & Fulfillment Metrics', index=False)

    # Best-Selling Products (Vertical Layout)
    best_selling_products_df = best_selling_products.reset_index()
    best_selling_products_df.columns = ['Product Name', 'Quantity Sold']
    best_selling_products_df.to_excel(writer, sheet_name='Best-Selling Products', index=False)

    # Revenue by Product (Vertical Layout)
    revenue_by_product_df = revenue_by_product.reset_index()
    revenue_by_product_df.columns = ['Product Name', 'Total Revenue']
    revenue_by_product_df.to_excel(writer, sheet_name='Revenue by Product', index=False)

    # Most Discounted Products (Vertical Layout)
    most_discounted_products_df = most_discounted_products.reset_index()
    most_discounted_products_df.columns = ['Product Name', 'Total Discount Given']
    most_discounted_products_df.to_excel(writer, sheet_name='Most Discounted Products', index=False)

    # Most Expensive Products Sold (Vertical Layout)
    most_expensive_products_df = most_expensive_products.reset_index()
    most_expensive_products_df.columns = ['Product Name', 'Max Price']
    most_expensive_products_df.to_excel(writer, sheet_name='Most Expensive Products Sold', index=False)

print("Metrics saved to 'financial_and_order_metrics.xlsx'")




