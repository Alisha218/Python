import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Load the CSV file
file_path = "orders_export_1.csv"  
excel_file = "Order_Analysis.xlsx"  

# Read CSV into DataFrame
df = pd.read_csv(file_path)

# Standardize city names (convert to lowercase and strip spaces)
df['Billing City'] = df['Billing City'].astype(str).str.strip().str.lower()

# Filter out rows where "Cancelled at" is not empty (meaning the order was canceled)
canceled_orders = df[df['Cancelled at'].notna()]

# Count the number of canceled orders per city
cancel_counts = canceled_orders.groupby('Billing City').size().reset_index(name='Cancel Count')

# Save to an existing Excel file without deleting other sheets
try:
    with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        cancel_counts.to_excel(writer, sheet_name="Cancelled Orders by City", index=False)
    print(f"Cancel counts saved to {excel_file} in 'Cancelled Orders by City' sheet.")
except FileNotFoundError:
    cancel_counts.to_excel(excel_file, sheet_name="Cancelled Orders by City", index=False)
    print(f"New file created: {excel_file} with 'Cancelled Orders by City' sheet.")

# Reopen the workbook after saving
wb = load_workbook(excel_file)

# Print available sheets to verify the correct name
print("Available sheets:", wb.sheetnames)

# Ensure correct sheet name
sheet_name = None
for name in wb.sheetnames:
    if "cancelled" in name.lower():
        sheet_name = name  # Assign the correct sheet name

if sheet_name is None:
    raise KeyError("Worksheet with 'Cancelled' in name not found!")

# Select the correct sheet
sheet = wb[sheet_name]

# Get the last row dynamically
last_row = sheet.max_row

# Define the range for your table dynamically
table_range = f"A1:B{last_row}"  # Adjust columns if necessary

# Create a table object
table = Table(displayName="CancelledOrdersByCity", ref=table_range)

# Apply Table Style
style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
table.tableStyleInfo = style

# Add the table to the sheet
sheet.add_table(table)

# Save and close
wb.save(excel_file)
wb.close()

print("Table formatting applied successfully!")

